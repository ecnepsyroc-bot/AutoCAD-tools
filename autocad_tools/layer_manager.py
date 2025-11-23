"""
Layer Manager Tool

Manages AutoCAD layer properties and configurations through Excel.
"""

from typing import Optional, List, Dict
import ezdxf
from openpyxl import Workbook
import pandas as pd
from .utils import (
    load_dxf, create_excel_workbook, style_header,
    auto_size_columns, add_border, save_excel, get_layer_names
)


class LayerManager:
    """
    Manages layer properties and configurations via Excel integration.
    """
    
    def __init__(self, dxf_path: str):
        """
        Initialize the layer manager.
        
        Args:
            dxf_path: Path to the DXF file
        """
        self.dxf_path = dxf_path
        self.doc = load_dxf(dxf_path)
        self.layers_data = []
        
    def extract_layer_info(self) -> List[Dict]:
        """
        Extract information about all layers.
        
        Returns:
            List of layer information dictionaries
        """
        self.layers_data = []
        
        for layer in self.doc.layers:
            layer_info = {
                'name': layer.dxf.name,
                'color': layer.dxf.color,
                'linetype': layer.dxf.linetype,
                'lineweight': getattr(layer.dxf, 'lineweight', -1),
                'on': not layer.is_off(),
                'frozen': layer.is_frozen(),
                'locked': layer.is_locked(),
                'plot': getattr(layer.dxf, 'plot', True)
            }
            self.layers_data.append(layer_info)
        
        return self.layers_data
    
    def to_excel(self, output_path: str):
        """
        Export layer configuration to Excel.
        
        Args:
            output_path: Path for output Excel file
        """
        if not self.layers_data:
            self.extract_layer_info()
        
        wb, ws = create_excel_workbook("Layers")
        
        # Headers
        headers = ['Name', 'Color', 'Linetype', 'Lineweight', 
                  'On', 'Frozen', 'Locked', 'Plot']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for idx, layer in enumerate(self.layers_data, start=2):
            ws.cell(row=idx, column=1, value=layer['name'])
            ws.cell(row=idx, column=2, value=layer['color'])
            ws.cell(row=idx, column=3, value=layer['linetype'])
            ws.cell(row=idx, column=4, value=layer['lineweight'])
            ws.cell(row=idx, column=5, value='Yes' if layer['on'] else 'No')
            ws.cell(row=idx, column=6, value='Yes' if layer['frozen'] else 'No')
            ws.cell(row=idx, column=7, value='Yes' if layer['locked'] else 'No')
            ws.cell(row=idx, column=8, value='Yes' if layer['plot'] else 'No')
        
        # Styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        last_row = len(self.layers_data) + 1
        add_border(ws, f"A1:H{last_row}")
        
        save_excel(wb, output_path)
    
    def from_excel(self, excel_path: str) -> List[Dict]:
        """
        Import layer configuration from Excel.
        
        Args:
            excel_path: Path to Excel file with layer configuration
            
        Returns:
            List of layer configurations
        """
        df = pd.read_excel(excel_path)
        
        layer_configs = []
        for _, row in df.iterrows():
            config = {
                'name': row.get('Name', ''),
                'color': int(row.get('Color', 7)),
                'linetype': row.get('Linetype', 'Continuous'),
                'lineweight': int(row.get('Lineweight', -1)),
                'on': str(row.get('On', 'Yes')).lower() in ['yes', 'true', '1'],
                'frozen': str(row.get('Frozen', 'No')).lower() in ['yes', 'true', '1'],
                'locked': str(row.get('Locked', 'No')).lower() in ['yes', 'true', '1'],
                'plot': str(row.get('Plot', 'Yes')).lower() in ['yes', 'true', '1']
            }
            layer_configs.append(config)
        
        return layer_configs
    
    def apply_layer_config(self, layer_configs: List[Dict]):
        """
        Apply layer configurations to the drawing.
        
        Args:
            layer_configs: List of layer configuration dictionaries
        """
        for config in layer_configs:
            layer_name = config['name']
            
            # Get or create layer
            if layer_name in self.doc.layers:
                layer = self.doc.layers.get(layer_name)
            else:
                layer = self.doc.layers.new(layer_name)
            
            # Apply properties
            layer.dxf.color = config['color']
            layer.dxf.linetype = config['linetype']
            
            if 'lineweight' in config:
                layer.dxf.lineweight = config['lineweight']
            
            # Set flags
            if not config['on']:
                layer.off()
            else:
                layer.on()
            
            if config['frozen']:
                layer.freeze()
            else:
                layer.thaw()
            
            if config['locked']:
                layer.lock()
            else:
                layer.unlock()
            
            if 'plot' in config:
                layer.dxf.plot = config['plot']
    
    def save_drawing(self, output_path: str):
        """
        Save the modified drawing.
        
        Args:
            output_path: Path for output DXF file
        """
        self.doc.saveas(output_path)
        print(f"Drawing saved: {output_path}")
    
    def get_layer_statistics(self) -> Dict:
        """
        Get statistics about layers and their usage.
        
        Returns:
            Dictionary with layer statistics
        """
        msp = self.doc.modelspace()
        layer_entity_counts = {}
        
        # Count entities per layer
        for entity in msp:
            layer = entity.dxf.layer
            layer_entity_counts[layer] = layer_entity_counts.get(layer, 0) + 1
        
        return {
            'total_layers': len(self.doc.layers),
            'entity_counts': layer_entity_counts,
            'empty_layers': [name for name in get_layer_names(self.doc) 
                           if name not in layer_entity_counts]
        }
    
    def create_layer_template(self, output_path: str):
        """
        Create an Excel template for layer configuration.
        
        Args:
            output_path: Path for template file
        """
        wb, ws = create_excel_workbook("Layer Template")
        
        # Headers with example data
        headers = ['Name', 'Color', 'Linetype', 'Lineweight', 
                  'On', 'Frozen', 'Locked', 'Plot']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Example layers
        examples = [
            ['0', 7, 'Continuous', -1, 'Yes', 'No', 'No', 'Yes'],
            ['WALLS', 1, 'Continuous', 50, 'Yes', 'No', 'No', 'Yes'],
            ['DOORS', 3, 'Continuous', 25, 'Yes', 'No', 'No', 'Yes'],
            ['WINDOWS', 5, 'Continuous', 25, 'Yes', 'No', 'No', 'Yes'],
            ['DIMENSIONS', 2, 'Continuous', -1, 'Yes', 'No', 'No', 'Yes'],
        ]
        
        for idx, example in enumerate(examples, start=2):
            for col, value in enumerate(example, start=1):
                ws.cell(row=idx, column=col, value=value)
        
        # Styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        add_border(ws, f"A1:H{len(examples) + 1}")
        
        save_excel(wb, output_path)

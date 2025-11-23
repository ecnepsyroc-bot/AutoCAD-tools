"""
Quantity Takeoff Tool

Performs automated measurements and quantity calculations from AutoCAD drawings.
"""

from typing import Optional, List, Dict
import ezdxf
from openpyxl import Workbook
import math
from .utils import (
    load_dxf, create_excel_workbook, style_header,
    auto_size_columns, add_border, save_excel, filter_by_layer
)


class QuantityTakeoff:
    """
    Automated quantity takeoff and measurement tool for AutoCAD drawings.
    """
    
    def __init__(self, dxf_path: str):
        """
        Initialize the quantity takeoff tool.
        
        Args:
            dxf_path: Path to the DXF file
        """
        self.dxf_path = dxf_path
        self.doc = load_dxf(dxf_path)
        self.msp = self.doc.modelspace()
        self.quantities = []
        
    def count_entities(self, entity_type: str, layer: Optional[str] = None) -> int:
        """
        Count entities of a specific type.
        
        Args:
            entity_type: Type of entity to count (e.g., 'LINE', 'CIRCLE', 'INSERT')
            layer: Layer to filter by
            
        Returns:
            Count of entities
        """
        entities = self.msp.query(entity_type)
        filtered = list(filter_by_layer(entities, layer))
        return len(filtered)
    
    def measure_lengths(self, layer: Optional[str] = None, 
                       entity_types: Optional[List[str]] = None) -> Dict:
        """
        Measure total lengths of linear entities.
        
        Args:
            layer: Layer to filter by
            entity_types: List of entity types to measure (default: ['LINE', 'LWPOLYLINE', 'POLYLINE'])
            
        Returns:
            Dictionary with length measurements
        """
        if entity_types is None:
            entity_types = ['LINE', 'LWPOLYLINE', 'POLYLINE']
        
        total_length = 0
        entity_counts = {}
        
        for entity_type in entity_types:
            entities = self.msp.query(entity_type)
            filtered = list(filter_by_layer(entities, layer))
            count = len(filtered)
            entity_counts[entity_type] = count
            
            if entity_type == 'LINE':
                for line in filtered:
                    length = math.sqrt(
                        (line.dxf.end.x - line.dxf.start.x)**2 +
                        (line.dxf.end.y - line.dxf.start.y)**2 +
                        (line.dxf.end.z - line.dxf.start.z)**2
                    )
                    total_length += length
            
            elif entity_type in ['LWPOLYLINE', 'POLYLINE']:
                for poly in filtered:
                    try:
                        # Get polyline length using ezdxf method
                        total_length += poly.get_length()
                    except:
                        # Fallback: calculate from points
                        points = list(poly.vertices())
                        for i in range(len(points) - 1):
                            p1 = points[i]
                            p2 = points[i + 1]
                            length = math.sqrt(
                                (p2[0] - p1[0])**2 + (p2[1] - p1[1])**2
                            )
                            total_length += length
        
        return {
            'total_length': total_length,
            'entity_counts': entity_counts,
            'layer': layer if layer else 'All layers'
        }
    
    def measure_areas(self, layer: Optional[str] = None) -> Dict:
        """
        Measure total areas of closed entities.
        
        Args:
            layer: Layer to filter by
            
        Returns:
            Dictionary with area measurements
        """
        total_area = 0
        entity_counts = {}
        
        # Measure circles
        circles = self.msp.query('CIRCLE')
        filtered_circles = list(filter_by_layer(circles, layer))
        entity_counts['CIRCLE'] = len(filtered_circles)
        
        for circle in filtered_circles:
            area = math.pi * circle.dxf.radius ** 2
            total_area += area
        
        # Measure polylines (if closed)
        for poly_type in ['LWPOLYLINE', 'POLYLINE']:
            polys = self.msp.query(poly_type)
            filtered_polys = list(filter_by_layer(polys, layer))
            closed_count = 0
            
            for poly in filtered_polys:
                if poly.is_closed:
                    closed_count += 1
                    try:
                        # Calculate area if possible
                        points = list(poly.vertices())
                        # Use shoelace formula for 2D polygon area
                        area = 0
                        for i in range(len(points)):
                            j = (i + 1) % len(points)
                            area += points[i][0] * points[j][1]
                            area -= points[j][0] * points[i][1]
                        total_area += abs(area) / 2
                    except:
                        pass
            
            entity_counts[f'{poly_type}_CLOSED'] = closed_count
        
        return {
            'total_area': total_area,
            'entity_counts': entity_counts,
            'layer': layer if layer else 'All layers'
        }
    
    def generate_takeoff_report(self, layer: Optional[str] = None) -> List[Dict]:
        """
        Generate a comprehensive quantity takeoff report.
        
        Args:
            layer: Layer to filter by
            
        Returns:
            List of quantity measurements
        """
        self.quantities = []
        
        # Count various entity types
        entity_types = ['LINE', 'CIRCLE', 'ARC', 'INSERT', 'TEXT', 
                       'LWPOLYLINE', 'POLYLINE', 'POINT']
        
        for etype in entity_types:
            count = self.count_entities(etype, layer)
            if count > 0:
                self.quantities.append({
                    'category': 'Count',
                    'item': etype,
                    'quantity': count,
                    'unit': 'EA',
                    'layer': layer if layer else 'All'
                })
        
        # Measure lengths
        length_data = self.measure_lengths(layer)
        if length_data['total_length'] > 0:
            self.quantities.append({
                'category': 'Length',
                'item': 'Total Linear Length',
                'quantity': round(length_data['total_length'], 2),
                'unit': 'units',
                'layer': layer if layer else 'All'
            })
        
        # Measure areas
        area_data = self.measure_areas(layer)
        if area_data['total_area'] > 0:
            self.quantities.append({
                'category': 'Area',
                'item': 'Total Enclosed Area',
                'quantity': round(area_data['total_area'], 2),
                'unit': 'sq units',
                'layer': layer if layer else 'All'
            })
        
        return self.quantities
    
    def to_excel(self, output_path: str):
        """
        Export quantity takeoff to Excel.
        
        Args:
            output_path: Path for output Excel file
        """
        if not self.quantities:
            raise ValueError("No quantities to export. Run generate_takeoff_report() first.")
        
        wb, ws = create_excel_workbook("Quantity Takeoff")
        
        # Headers
        headers = ['Index', 'Category', 'Item', 'Quantity', 'Unit', 'Layer']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for idx, qty in enumerate(self.quantities, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=qty['category'])
            ws.cell(row=idx, column=3, value=qty['item'])
            ws.cell(row=idx, column=4, value=qty['quantity'])
            ws.cell(row=idx, column=5, value=qty['unit'])
            ws.cell(row=idx, column=6, value=qty['layer'])
        
        # Styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        last_row = len(self.quantities) + 1
        add_border(ws, f"A1:F{last_row}")
        
        save_excel(wb, output_path)

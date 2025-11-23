"""
Measurement and Analysis Tools

Provides intelligent measurement and analysis capabilities with Excel reporting.
"""

from typing import List, Dict, Optional, Tuple
import math
import ezdxf
from openpyxl import Workbook
from .utils import (
    load_dxf, create_excel_workbook, style_header,
    auto_size_columns, add_border, save_excel, filter_by_layer,
    calculate_distance
)


class MeasurementTools:
    """
    Advanced measurement and analysis tools for AutoCAD drawings.
    """
    
    def __init__(self, dxf_path: str):
        """
        Initialize measurement tools.
        
        Args:
            dxf_path: Path to the DXF file
        """
        self.dxf_path = dxf_path
        self.doc = load_dxf(dxf_path)
        self.msp = self.doc.modelspace()
        self.measurements = []
        
    def measure_distances(self, point_pairs: List[Tuple[Tuple, Tuple]]) -> List[Dict]:
        """
        Measure distances between point pairs.
        
        Args:
            point_pairs: List of point pair tuples ((x1,y1,z1), (x2,y2,z2))
            
        Returns:
            List of distance measurements
        """
        distances = []
        
        for idx, (p1, p2) in enumerate(point_pairs):
            distance = calculate_distance(p1, p2)
            distances.append({
                'index': idx + 1,
                'point1': p1,
                'point2': p2,
                'distance': distance,
                '2d_distance': math.sqrt((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)
            })
        
        return distances
    
    def analyze_polyline(self, layer: Optional[str] = None) -> List[Dict]:
        """
        Analyze polyline properties.
        
        Args:
            layer: Layer to filter by
            
        Returns:
            List of polyline analysis results
        """
        results = []
        
        for poly_type in ['LWPOLYLINE', 'POLYLINE']:
            entities = self.msp.query(poly_type)
            filtered = list(filter_by_layer(entities, layer))
            
            for idx, poly in enumerate(filtered):
                try:
                    length = poly.get_length()
                    is_closed = poly.is_closed
                    
                    vertices = list(poly.vertices())
                    vertex_count = len(vertices)
                    
                    # Calculate bounding box
                    x_coords = [v[0] for v in vertices]
                    y_coords = [v[1] for v in vertices]
                    
                    bbox = {
                        'min_x': min(x_coords),
                        'max_x': max(x_coords),
                        'min_y': min(y_coords),
                        'max_y': max(y_coords),
                        'width': max(x_coords) - min(x_coords),
                        'height': max(y_coords) - min(y_coords)
                    }
                    
                    result = {
                        'index': idx + 1,
                        'type': poly_type,
                        'layer': poly.dxf.layer,
                        'length': length,
                        'vertices': vertex_count,
                        'closed': is_closed,
                        **bbox
                    }
                    
                    # Calculate area if closed
                    if is_closed:
                        try:
                            area = 0
                            for i in range(len(vertices)):
                                j = (i + 1) % len(vertices)
                                area += vertices[i][0] * vertices[j][1]
                                area -= vertices[j][0] * vertices[i][1]
                            result['area'] = abs(area) / 2
                        except:
                            result['area'] = None
                    else:
                        result['area'] = None
                    
                    results.append(result)
                    
                except Exception as e:
                    print(f"Error analyzing polyline: {e}")
        
        return results
    
    def analyze_circles(self, layer: Optional[str] = None) -> List[Dict]:
        """
        Analyze circle properties.
        
        Args:
            layer: Layer to filter by
            
        Returns:
            List of circle analysis results
        """
        circles = self.msp.query('CIRCLE')
        filtered = list(filter_by_layer(circles, layer))
        
        results = []
        for idx, circle in enumerate(filtered):
            results.append({
                'index': idx + 1,
                'layer': circle.dxf.layer,
                'center_x': circle.dxf.center.x,
                'center_y': circle.dxf.center.y,
                'center_z': circle.dxf.center.z,
                'radius': circle.dxf.radius,
                'diameter': circle.dxf.radius * 2,
                'circumference': 2 * math.pi * circle.dxf.radius,
                'area': math.pi * circle.dxf.radius ** 2
            })
        
        return results
    
    def calculate_statistics(self, measurements: List[Dict], value_key: str) -> Dict:
        """
        Calculate statistical summary of measurements.
        
        Args:
            measurements: List of measurement dictionaries
            value_key: Key to calculate statistics for
            
        Returns:
            Dictionary with statistical values
        """
        if not measurements:
            return {}
        
        values = [m[value_key] for m in measurements if value_key in m and m[value_key] is not None]
        
        if not values:
            return {}
        
        return {
            'count': len(values),
            'min': min(values),
            'max': max(values),
            'mean': sum(values) / len(values),
            'total': sum(values),
            'range': max(values) - min(values)
        }
    
    def comprehensive_analysis(self, layer: Optional[str] = None) -> Dict:
        """
        Perform comprehensive analysis of the drawing.
        
        Args:
            layer: Layer to filter by
            
        Returns:
            Dictionary with comprehensive analysis results
        """
        analysis = {
            'polylines': self.analyze_polyline(layer),
            'circles': self.analyze_circles(layer)
        }
        
        # Calculate statistics
        if analysis['polylines']:
            analysis['polyline_stats'] = {
                'length': self.calculate_statistics(analysis['polylines'], 'length'),
                'area': self.calculate_statistics(analysis['polylines'], 'area')
            }
        
        if analysis['circles']:
            analysis['circle_stats'] = {
                'radius': self.calculate_statistics(analysis['circles'], 'radius'),
                'area': self.calculate_statistics(analysis['circles'], 'area')
            }
        
        return analysis
    
    def export_analysis_to_excel(self, output_path: str, 
                                 layer: Optional[str] = None):
        """
        Export comprehensive analysis to Excel.
        
        Args:
            output_path: Path for output Excel file
            layer: Layer to filter by
        """
        analysis = self.comprehensive_analysis(layer)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Polylines sheet
        if analysis['polylines']:
            ws_poly = wb.create_sheet("Polylines")
            headers = ['Index', 'Type', 'Layer', 'Length', 'Vertices', 
                      'Closed', 'Area', 'Width', 'Height']
            
            for col, header in enumerate(headers, start=1):
                ws_poly.cell(row=1, column=col, value=header)
            
            for idx, poly in enumerate(analysis['polylines'], start=2):
                ws_poly.cell(row=idx, column=1, value=poly['index'])
                ws_poly.cell(row=idx, column=2, value=poly['type'])
                ws_poly.cell(row=idx, column=3, value=poly['layer'])
                ws_poly.cell(row=idx, column=4, value=round(poly['length'], 4))
                ws_poly.cell(row=idx, column=5, value=poly['vertices'])
                ws_poly.cell(row=idx, column=6, value='Yes' if poly['closed'] else 'No')
                ws_poly.cell(row=idx, column=7, value=round(poly['area'], 4) if poly['area'] else '')
                ws_poly.cell(row=idx, column=8, value=round(poly['width'], 4))
                ws_poly.cell(row=idx, column=9, value=round(poly['height'], 4))
            
            style_header(ws_poly, row=1)
            auto_size_columns(ws_poly)
            add_border(ws_poly, f"A1:I{len(analysis['polylines']) + 1}")
        
        # Circles sheet
        if analysis['circles']:
            ws_circ = wb.create_sheet("Circles")
            headers = ['Index', 'Layer', 'Center X', 'Center Y', 'Center Z',
                      'Radius', 'Diameter', 'Circumference', 'Area']
            
            for col, header in enumerate(headers, start=1):
                ws_circ.cell(row=1, column=col, value=header)
            
            for idx, circle in enumerate(analysis['circles'], start=2):
                ws_circ.cell(row=idx, column=1, value=circle['index'])
                ws_circ.cell(row=idx, column=2, value=circle['layer'])
                ws_circ.cell(row=idx, column=3, value=round(circle['center_x'], 4))
                ws_circ.cell(row=idx, column=4, value=round(circle['center_y'], 4))
                ws_circ.cell(row=idx, column=5, value=round(circle['center_z'], 4))
                ws_circ.cell(row=idx, column=6, value=round(circle['radius'], 4))
                ws_circ.cell(row=idx, column=7, value=round(circle['diameter'], 4))
                ws_circ.cell(row=idx, column=8, value=round(circle['circumference'], 4))
                ws_circ.cell(row=idx, column=9, value=round(circle['area'], 4))
            
            style_header(ws_circ, row=1)
            auto_size_columns(ws_circ)
            add_border(ws_circ, f"A1:I{len(analysis['circles']) + 1}")
        
        # Statistics sheet
        ws_stats = wb.create_sheet("Statistics")
        row = 1
        
        if 'polyline_stats' in analysis:
            ws_stats.cell(row=row, column=1, value="Polyline Statistics")
            ws_stats.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            for stat_type, stats in analysis['polyline_stats'].items():
                if stats:
                    ws_stats.cell(row=row, column=1, value=f"{stat_type.title()}:")
                    ws_stats.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    
                    for key, value in stats.items():
                        ws_stats.cell(row=row, column=1, value=key.title())
                        ws_stats.cell(row=row, column=2, value=round(value, 4))
                        row += 1
                    row += 1
        
        if 'circle_stats' in analysis:
            ws_stats.cell(row=row, column=1, value="Circle Statistics")
            ws_stats.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            for stat_type, stats in analysis['circle_stats'].items():
                if stats:
                    ws_stats.cell(row=row, column=1, value=f"{stat_type.title()}:")
                    ws_stats.cell(row=row, column=1).font = Font(bold=True)
                    row += 1
                    
                    for key, value in stats.items():
                        ws_stats.cell(row=row, column=1, value=key.title())
                        ws_stats.cell(row=row, column=2, value=round(value, 4))
                        row += 1
        
        auto_size_columns(ws_stats)
        
        # Import Font at top of function if not already imported
        from openpyxl.styles import Font
        
        save_excel(wb, output_path)

"""
Block Attribute Extractor Tool

Extracts block information and attributes from AutoCAD drawings.
"""

from typing import Optional, List, Dict
import ezdxf
from openpyxl import Workbook
from .utils import (
    load_dxf, create_excel_workbook, style_header,
    auto_size_columns, add_border, save_excel, filter_by_layer
)


class BlockAttributeExtractor:
    """
    Extracts block information and attributes from AutoCAD DXF files.
    """
    
    def __init__(self, dxf_path: str):
        """
        Initialize the block attribute extractor.
        
        Args:
            dxf_path: Path to the DXF file
        """
        self.dxf_path = dxf_path
        self.doc = load_dxf(dxf_path)
        self.msp = self.doc.modelspace()
        self.blocks = []
        
    def extract_all_blocks(self, layer: Optional[str] = None) -> List[Dict]:
        """
        Extract all block insertions with their attributes.
        
        Args:
            layer: Layer name to filter by
            
        Returns:
            List of block dictionaries
        """
        self.blocks = []
        entities = self.msp.query('INSERT')
        
        for insert in filter_by_layer(entities, layer):
            block_data = {
                'name': insert.dxf.name,
                'x': insert.dxf.insert.x,
                'y': insert.dxf.insert.y,
                'z': insert.dxf.insert.z,
                'layer': insert.dxf.layer,
                'rotation': getattr(insert.dxf, 'rotation', 0),
                'x_scale': getattr(insert.dxf, 'xscale', 1),
                'y_scale': getattr(insert.dxf, 'yscale', 1),
                'z_scale': getattr(insert.dxf, 'zscale', 1),
                'attributes': {}
            }
            
            # Extract attributes if present
            if insert.has_attrib:
                for attrib in insert.attribs:
                    block_data['attributes'][attrib.dxf.tag] = attrib.dxf.text
            
            self.blocks.append(block_data)
        
        return self.blocks
    
    def extract_by_name(self, block_name: str, layer: Optional[str] = None) -> List[Dict]:
        """
        Extract blocks by name.
        
        Args:
            block_name: Name of the block to extract
            layer: Layer name to filter by
            
        Returns:
            List of matching block dictionaries
        """
        self.extract_all_blocks(layer)
        return [b for b in self.blocks if b['name'] == block_name]
    
    def to_excel(self, output_path: str, separate_attributes: bool = True):
        """
        Export extracted blocks to Excel.
        
        Args:
            output_path: Path for the output Excel file
            separate_attributes: Create separate columns for each attribute
        """
        if not self.blocks:
            raise ValueError("No blocks to export. Run extract_all_blocks() first.")
        
        wb, ws = create_excel_workbook("Blocks")
        
        # Collect all unique attribute keys
        all_attr_keys = set()
        if separate_attributes:
            for block in self.blocks:
                all_attr_keys.update(block['attributes'].keys())
        
        # Create headers
        headers = ['Index', 'Block Name', 'X', 'Y', 'Z', 'Layer', 
                  'Rotation', 'X Scale', 'Y Scale', 'Z Scale']
        
        if separate_attributes and all_attr_keys:
            headers.extend(sorted(all_attr_keys))
        elif not separate_attributes:
            headers.append('Attributes')
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for idx, block in enumerate(self.blocks, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=block['name'])
            ws.cell(row=idx, column=3, value=round(block['x'], 4))
            ws.cell(row=idx, column=4, value=round(block['y'], 4))
            ws.cell(row=idx, column=5, value=round(block['z'], 4))
            ws.cell(row=idx, column=6, value=block['layer'])
            ws.cell(row=idx, column=7, value=round(block['rotation'], 2))
            ws.cell(row=idx, column=8, value=round(block['x_scale'], 2))
            ws.cell(row=idx, column=9, value=round(block['y_scale'], 2))
            ws.cell(row=idx, column=10, value=round(block['z_scale'], 2))
            
            if separate_attributes and all_attr_keys:
                col_offset = 11
                for attr_key in sorted(all_attr_keys):
                    value = block['attributes'].get(attr_key, '')
                    ws.cell(row=idx, column=col_offset, value=value)
                    col_offset += 1
            elif not separate_attributes:
                attr_str = ', '.join([f"{k}: {v}" for k, v in block['attributes'].items()])
                ws.cell(row=idx, column=11, value=attr_str)
        
        # Apply styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        # Add borders
        last_row = len(self.blocks) + 1
        last_col = len(headers)
        add_border(ws, f"A1:{chr(64 + last_col)}{last_row}")
        
        save_excel(wb, output_path)
    
    def get_summary(self) -> Dict:
        """
        Get summary statistics of extracted blocks.
        
        Returns:
            Dictionary with summary information
        """
        if not self.blocks:
            return {'count': 0}
        
        block_names = {}
        layers = {}
        attribute_count = 0
        
        for block in self.blocks:
            name = block['name']
            layer = block['layer']
            
            block_names[name] = block_names.get(name, 0) + 1
            layers[layer] = layers.get(layer, 0) + 1
            attribute_count += len(block['attributes'])
        
        return {
            'count': len(self.blocks),
            'unique_blocks': len(block_names),
            'block_types': block_names,
            'layers': layers,
            'total_attributes': attribute_count,
            'avg_attributes_per_block': attribute_count / len(self.blocks) if self.blocks else 0
        }
    
    def filter_blocks(self, name: Optional[str] = None, 
                     layer: Optional[str] = None) -> List[Dict]:
        """
        Filter extracted blocks by criteria.
        
        Args:
            name: Block name to filter by
            layer: Layer name to filter by
            
        Returns:
            Filtered list of blocks
        """
        filtered = self.blocks
        
        if name:
            filtered = [b for b in filtered if b['name'] == name]
        if layer:
            filtered = [b for b in filtered if b['layer'] == layer]
        
        return filtered

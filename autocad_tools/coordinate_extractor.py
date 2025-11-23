"""
Coordinate Extractor Tool

Extracts point coordinates from AutoCAD drawings and exports to Excel.
"""

from typing import Optional, List, Dict
import ezdxf
from openpyxl import Workbook
from .utils import (
    load_dxf, create_excel_workbook, style_header, 
    auto_size_columns, add_border, save_excel, filter_by_layer
)


class CoordinateExtractor:
    """
    Extracts point coordinates from AutoCAD DXF files and exports to Excel.
    """
    
    def __init__(self, dxf_path: str):
        """
        Initialize the coordinate extractor.
        
        Args:
            dxf_path: Path to the DXF file
        """
        self.dxf_path = dxf_path
        self.doc = load_dxf(dxf_path)
        self.msp = self.doc.modelspace()
        self.points = []
        
    def extract_points(self, layer: Optional[str] = None, 
                      entity_type: str = 'POINT') -> List[Dict]:
        """
        Extract point coordinates from the drawing.
        
        Args:
            layer: Layer name to filter by (None for all layers)
            entity_type: Entity type to extract ('POINT', 'INSERT', 'CIRCLE', 'LINE')
            
        Returns:
            List of point dictionaries with coordinates and metadata
        """
        self.points = []
        
        if entity_type == 'POINT':
            entities = self.msp.query('POINT')
            for point in filter_by_layer(entities, layer):
                self.points.append({
                    'type': 'POINT',
                    'x': point.dxf.location.x,
                    'y': point.dxf.location.y,
                    'z': point.dxf.location.z,
                    'layer': point.dxf.layer
                })
                
        elif entity_type == 'INSERT':
            entities = self.msp.query('INSERT')
            for insert in filter_by_layer(entities, layer):
                self.points.append({
                    'type': 'INSERT',
                    'x': insert.dxf.insert.x,
                    'y': insert.dxf.insert.y,
                    'z': insert.dxf.insert.z,
                    'layer': insert.dxf.layer,
                    'name': insert.dxf.name
                })
                
        elif entity_type == 'CIRCLE':
            entities = self.msp.query('CIRCLE')
            for circle in filter_by_layer(entities, layer):
                self.points.append({
                    'type': 'CIRCLE',
                    'x': circle.dxf.center.x,
                    'y': circle.dxf.center.y,
                    'z': circle.dxf.center.z,
                    'layer': circle.dxf.layer,
                    'radius': circle.dxf.radius
                })
                
        elif entity_type == 'LINE':
            entities = self.msp.query('LINE')
            for line in filter_by_layer(entities, layer):
                # Add both start and end points
                self.points.append({
                    'type': 'LINE_START',
                    'x': line.dxf.start.x,
                    'y': line.dxf.start.y,
                    'z': line.dxf.start.z,
                    'layer': line.dxf.layer
                })
                self.points.append({
                    'type': 'LINE_END',
                    'x': line.dxf.end.x,
                    'y': line.dxf.end.y,
                    'z': line.dxf.end.z,
                    'layer': line.dxf.layer
                })
        
        return self.points
    
    def extract_all_points(self, layer: Optional[str] = None) -> List[Dict]:
        """
        Extract all point-like entities from the drawing.
        
        Args:
            layer: Layer name to filter by
            
        Returns:
            List of all point dictionaries
        """
        all_points = []
        for entity_type in ['POINT', 'INSERT', 'CIRCLE']:
            all_points.extend(self.extract_points(layer, entity_type))
        return all_points
    
    def to_excel(self, output_path: str, include_metadata: bool = True):
        """
        Export extracted points to Excel.
        
        Args:
            output_path: Path for the output Excel file
            include_metadata: Whether to include entity metadata
        """
        if not self.points:
            raise ValueError("No points to export. Run extract_points() first.")
        
        wb, ws = create_excel_workbook("Coordinates")
        
        # Create headers
        headers = ['Index', 'Type', 'X', 'Y', 'Z', 'Layer']
        if include_metadata:
            # Check if we have additional fields
            extra_fields = set()
            for point in self.points:
                for key in point.keys():
                    if key not in ['type', 'x', 'y', 'z', 'layer']:
                        extra_fields.add(key)
            headers.extend(sorted(extra_fields))
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for idx, point in enumerate(self.points, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=point.get('type', ''))
            ws.cell(row=idx, column=3, value=round(point['x'], 4))
            ws.cell(row=idx, column=4, value=round(point['y'], 4))
            ws.cell(row=idx, column=5, value=round(point['z'], 4))
            ws.cell(row=idx, column=6, value=point.get('layer', ''))
            
            if include_metadata:
                col_offset = 7
                for field in sorted(extra_fields):
                    if field in point:
                        ws.cell(row=idx, column=col_offset, 
                               value=point[field])
                    col_offset += 1
        
        # Apply styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        # Add borders
        last_row = len(self.points) + 1
        last_col = len(headers)
        add_border(ws, f"A1:{chr(64 + last_col)}{last_row}")
        
        save_excel(wb, output_path)
        
    def get_summary(self) -> Dict:
        """
        Get summary statistics of extracted points.
        
        Returns:
            Dictionary with summary information
        """
        if not self.points:
            return {'count': 0}
        
        types = {}
        layers = {}
        
        for point in self.points:
            point_type = point.get('type', 'UNKNOWN')
            layer = point.get('layer', 'UNKNOWN')
            
            types[point_type] = types.get(point_type, 0) + 1
            layers[layer] = layers.get(layer, 0) + 1
        
        return {
            'count': len(self.points),
            'types': types,
            'layers': layers,
            'x_range': (min(p['x'] for p in self.points), 
                       max(p['x'] for p in self.points)),
            'y_range': (min(p['y'] for p in self.points), 
                       max(p['y'] for p in self.points)),
            'z_range': (min(p['z'] for p in self.points), 
                       max(p['z'] for p in self.points))
        }

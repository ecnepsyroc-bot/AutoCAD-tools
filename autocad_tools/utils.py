"""
Utility functions for AutoCAD-Excel integration.
"""

import os
from typing import Optional, List, Tuple
import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def load_dxf(filepath: str):
    """
    Load a DXF file.
    
    Args:
        filepath: Path to the DXF file
        
    Returns:
        ezdxf document object
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    return ezdxf.readfile(filepath)


def create_excel_workbook(title: str = "AutoCAD Export") -> Tuple[Workbook, any]:
    """
    Create a new Excel workbook with styling.
    
    Args:
        title: Title for the first worksheet
        
    Returns:
        Tuple of (workbook, worksheet)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def style_header(ws, row: int = 1, columns: Optional[List[str]] = None):
    """
    Apply header styling to Excel worksheet.
    
    Args:
        ws: Worksheet object
        row: Row number for header
        columns: List of column letters to style
    """
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if columns is None:
        # Style all cells in the row that have values
        for cell in ws[row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    else:
        for col in columns:
            cell = ws[f"{col}{row}"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


def auto_size_columns(ws, min_width: int = 10, max_width: int = 50):
    """
    Auto-size columns based on content.
    
    Args:
        ws: Worksheet object
        min_width: Minimum column width
        max_width: Maximum column width
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_border(ws, cell_range: str):
    """
    Add borders to a cell range.
    
    Args:
        ws: Worksheet object
        cell_range: Range like "A1:D10"
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border


def get_layer_names(doc) -> List[str]:
    """
    Get all layer names from a DXF document.
    
    Args:
        doc: ezdxf document object
        
    Returns:
        List of layer names
    """
    return [layer.dxf.name for layer in doc.layers]


def filter_by_layer(entities, layer_name: Optional[str] = None):
    """
    Filter entities by layer name.
    
    Args:
        entities: Iterable of entities
        layer_name: Layer name to filter by, None for all
        
    Returns:
        Filtered entities
    """
    if layer_name is None:
        return entities
    return [e for e in entities if e.dxf.layer == layer_name]


def format_coordinate(value: float, precision: int = 3) -> str:
    """
    Format a coordinate value.
    
    Args:
        value: Coordinate value
        precision: Decimal places
        
    Returns:
        Formatted string
    """
    return f"{value:.{precision}f}"


def calculate_distance(p1: Tuple[float, float, float], 
                       p2: Tuple[float, float, float]) -> float:
    """
    Calculate 3D distance between two points.
    
    Args:
        p1: First point (x, y, z)
        p2: Second point (x, y, z)
        
    Returns:
        Distance
    """
    return ((p2[0] - p1[0])**2 + (p2[1] - p1[1])**2 + (p2[2] - p1[2])**2)**0.5


def save_excel(wb: Workbook, filepath: str):
    """
    Save Excel workbook with error handling.
    
    Args:
        wb: Workbook object
        filepath: Output file path
    """
    try:
        wb.save(filepath)
        print(f"Successfully saved: {filepath}")
    except PermissionError:
        raise PermissionError(f"Cannot save file (file may be open): {filepath}")
    except Exception as e:
        raise Exception(f"Error saving file: {str(e)}")

"""
Batch Processor Tool

Executes batch operations on multiple AutoCAD drawings based on Excel input.
"""

from typing import List, Dict, Optional
import os
import pandas as pd
from openpyxl import Workbook
import ezdxf
from .utils import (
    load_dxf, create_excel_workbook, style_header,
    auto_size_columns, add_border, save_excel
)


class BatchProcessor:
    """
    Executes batch operations on multiple drawings from Excel configurations.
    """
    
    def __init__(self):
        """Initialize the batch processor."""
        self.results = []
        
    def create_batch_template(self, output_path: str):
        """
        Create an Excel template for batch processing.
        
        Args:
            output_path: Path for template file
        """
        wb, ws = create_excel_workbook("Batch Template")
        
        # Headers
        headers = ['Input File', 'Output File', 'Operation', 
                  'Layer Filter', 'Parameters', 'Notes']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Example operations
        examples = [
            ['drawing1.dxf', 'output1.dxf', 'extract_blocks', 'ALL', '', 'Extract all blocks'],
            ['drawing2.dxf', 'output2.dxf', 'extract_points', 'POINTS', '', 'Extract points layer'],
            ['drawing3.dxf', 'output3.dxf', 'quantity_takeoff', '', '', 'Generate quantities'],
        ]
        
        for idx, example in enumerate(examples, start=2):
            for col, value in enumerate(example, start=1):
                ws.cell(row=idx, column=col, value=value)
        
        # Styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        add_border(ws, f"A1:F{len(examples) + 1}")
        
        save_excel(wb, output_path)
    
    def load_batch_config(self, excel_path: str) -> List[Dict]:
        """
        Load batch processing configuration from Excel.
        
        Args:
            excel_path: Path to Excel configuration file
            
        Returns:
            List of batch operation configurations
        """
        df = pd.read_excel(excel_path)
        
        configs = []
        for _, row in df.iterrows():
            config = {
                'input_file': row.get('Input File', ''),
                'output_file': row.get('Output File', ''),
                'operation': row.get('Operation', ''),
                'layer_filter': row.get('Layer Filter', None),
                'parameters': row.get('Parameters', ''),
                'notes': row.get('Notes', '')
            }
            configs.append(config)
        
        return configs
    
    def process_drawing(self, config: Dict) -> Dict:
        """
        Process a single drawing based on configuration.
        
        Args:
            config: Configuration dictionary
            
        Returns:
            Result dictionary with status and message
        """
        result = {
            'input_file': config['input_file'],
            'operation': config['operation'],
            'status': 'Unknown',
            'message': ''
        }
        
        try:
            # Check if input file exists
            if not os.path.exists(config['input_file']):
                result['status'] = 'Error'
                result['message'] = f"File not found: {config['input_file']}"
                return result
            
            # Load the drawing
            doc = load_dxf(config['input_file'])
            
            # Execute operation
            operation = config['operation'].lower()
            
            if operation == 'extract_blocks':
                from .block_extractor import BlockAttributeExtractor
                extractor = BlockAttributeExtractor(config['input_file'])
                layer = None if config['layer_filter'] == 'ALL' else config['layer_filter']
                extractor.extract_all_blocks(layer)
                extractor.to_excel(config['output_file'])
                result['status'] = 'Success'
                result['message'] = f"Extracted {len(extractor.blocks)} blocks"
                
            elif operation == 'extract_points':
                from .coordinate_extractor import CoordinateExtractor
                extractor = CoordinateExtractor(config['input_file'])
                layer = None if config['layer_filter'] == 'ALL' else config['layer_filter']
                extractor.extract_all_points(layer)
                extractor.to_excel(config['output_file'])
                result['status'] = 'Success'
                result['message'] = f"Extracted {len(extractor.points)} points"
                
            elif operation == 'quantity_takeoff':
                from .quantity_takeoff import QuantityTakeoff
                takeoff = QuantityTakeoff(config['input_file'])
                layer = None if config['layer_filter'] == 'ALL' else config['layer_filter']
                takeoff.generate_takeoff_report(layer)
                takeoff.to_excel(config['output_file'])
                result['status'] = 'Success'
                result['message'] = f"Generated {len(takeoff.quantities)} quantity items"
                
            elif operation == 'layer_export':
                from .layer_manager import LayerManager
                manager = LayerManager(config['input_file'])
                manager.extract_layer_info()
                manager.to_excel(config['output_file'])
                result['status'] = 'Success'
                result['message'] = f"Exported {len(manager.layers_data)} layers"
                
            else:
                result['status'] = 'Error'
                result['message'] = f"Unknown operation: {operation}"
        
        except Exception as e:
            result['status'] = 'Error'
            result['message'] = str(e)
        
        return result
    
    def process_batch(self, excel_path: str, base_path: Optional[str] = None) -> List[Dict]:
        """
        Process multiple drawings from Excel configuration.
        
        Args:
            excel_path: Path to Excel configuration file
            base_path: Base directory for relative paths (default: directory of excel_path)
            
        Returns:
            List of processing results
        """
        if base_path is None:
            base_path = os.path.dirname(os.path.abspath(excel_path))
        
        configs = self.load_batch_config(excel_path)
        self.results = []
        
        for config in configs:
            # Resolve relative paths
            if not os.path.isabs(config['input_file']):
                config['input_file'] = os.path.join(base_path, config['input_file'])
            if not os.path.isabs(config['output_file']):
                config['output_file'] = os.path.join(base_path, config['output_file'])
            
            # Process the drawing
            result = self.process_drawing(config)
            self.results.append(result)
            
            # Print progress
            print(f"[{result['status']}] {config['input_file']} - {result['message']}")
        
        return self.results
    
    def export_results(self, output_path: str):
        """
        Export batch processing results to Excel.
        
        Args:
            output_path: Path for output Excel file
        """
        if not self.results:
            raise ValueError("No results to export. Run process_batch() first.")
        
        wb, ws = create_excel_workbook("Batch Results")
        
        # Headers
        headers = ['Index', 'Input File', 'Operation', 'Status', 'Message']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for idx, result in enumerate(self.results, start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=result['input_file'])
            ws.cell(row=idx, column=3, value=result['operation'])
            ws.cell(row=idx, column=4, value=result['status'])
            ws.cell(row=idx, column=5, value=result['message'])
        
        # Styling
        style_header(ws, row=1)
        auto_size_columns(ws)
        
        last_row = len(self.results) + 1
        add_border(ws, f"A1:E{last_row}")
        
        save_excel(wb, output_path)
    
    def get_summary(self) -> Dict:
        """
        Get summary of batch processing results.
        
        Returns:
            Summary dictionary
        """
        if not self.results:
            return {'total': 0}
        
        success_count = sum(1 for r in self.results if r['status'] == 'Success')
        error_count = sum(1 for r in self.results if r['status'] == 'Error')
        
        return {
            'total': len(self.results),
            'success': success_count,
            'errors': error_count,
            'success_rate': f"{(success_count / len(self.results) * 100):.1f}%"
        }
